import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

def generate_strict_template(file_path: str):
    """
    Generates the strict API Input Template with 3 sheets:
    1. apis (Main data)
    2. environments (Variables)
    3. rules (Configuration)
    """
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: apis ---
    ws_apis = wb.active
    ws_apis.title = "apis"
    
    headers = [
        "api_group", "api_name", "http_method", "base_url_var", "endpoint", 
        "body_type", "request_payload", "headers", "expected_status", 
        "expected_response", "extract_variable", "description"
    ]
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_apis.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    
    # Freeze header
    ws_apis.freeze_panes = "A2"
    
    # Data Validation: http_method
    dv_method = DataValidation(type="list", formula1='"GET,POST,PUT,DELETE,PATCH"', allow_blank=False)
    ws_apis.add_data_validation(dv_method)
    dv_method.add("C2:C1000") # Apply to http_method column
    
    # Data Validation: body_type
    dv_body = DataValidation(type="list", formula1='"json,urlencoded,none"', allow_blank=False)
    ws_apis.add_data_validation(dv_body)
    dv_body.add("F2:F1000") # Apply to body_type column

    # Example Data
    examples = [
        ["Auth", "getAuthToken", "POST", "basic_url", "/getAuthToken", "urlencoded", "", "Content-Type: application/x-www-form-urlencoded", 200, '{"token": "..."}', "token:response.token", "Login to get token"],
        ["Users", "createUser", "POST", "basic_url", "/users", "json", '{"name": "John"}', "Content-Type: application/json", 201, '{"id": 123}', "user_id:response.id", "Create a new user"],
        ["Users", "getUser", "GET", "basic_url", "/users/{{user_id}}", "none", "", "Authorization: Bearer {{token}}", 200, '{"name": "John"}', "", "Get user details"]
    ]
    
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_apis.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust columns
    for col in ws_apis.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_apis.column_dimensions[column].width = max_length + 2


    # --- Sheet 2: environments ---
    ws_env = wb.create_sheet("environments")
    env_headers = ["variable", "value", "description"]
    for col_num, header in enumerate(env_headers, 1):
        cell = ws_env.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    ws_env.append(["basic_url", "https://api.example.com", "Base URL for the API"])
    ws_env.append(["token", "", "Auth token (auto-filled usually)"])
    ws_env.column_dimensions["A"].width = 20
    ws_env.column_dimensions["B"].width = 40
    ws_env.column_dimensions["C"].width = 30


    # --- Sheet 3: rules ---
    ws_rules = wb.create_sheet("rules")
    rule_headers = ["rule", "value", "description"]
    for col_num, header in enumerate(rule_headers, 1):
        cell = ws_rules.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        
    rules_data = [
        ["strict_validation", "true", "Fail if any row has errors"],
        ["fail_on_invalid_json", "true", "Fail if payload is not valid JSON"],
        ["auto_token_detection", "true", "Auto-extract tokens from 'auth' endpoints"]
    ]
    
    for r in rules_data:
        ws_rules.append(r)
        
    ws_rules.column_dimensions["A"].width = 25
    ws_rules.column_dimensions["B"].width = 15
    ws_rules.column_dimensions["C"].width = 50

    wb.save(file_path)
    return file_path
